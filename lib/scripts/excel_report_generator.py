#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Report Generator for As-Built Installation Reports
Generates professional Excel reports with multiple phases, advanced analytics, and conditional formatting
"""

import sys
import json
import io
from datetime import datetime, timedelta
from collections import defaultdict

# Ensure UTF-8 encoding for stdout on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, BarChart3D, PieChart, PieChart3D, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.styles import colors

# ═════════════════════════════════════════════════════════════════
# CONSTANTES DE CORES E ESTILOS
# ═════════════════════════════════════════════════════════════════

COLORS = {
    'header': '1F4E78',           # Azul profissional
    'subheader': '4472C4',        # Azul claro
    'success': '70AD47',          # Verde
    'warning': 'FFC7CE',          # Vermelho claro (alertas)
    'caution': 'FFEB9C',          # Amarelo (aviso)
    'danger': 'FF0000',           # Vermelho (erro)
    'info': 'D9E1F2',             # Azul muito claro
    'white': 'FFFFFF',
    'light_gray': 'F2F2F2',
    'dark_gray': '666666',
    # Paleta sincronizada com gráficos Excel (baseada em observação real)
    'excel_blue': '5B9BD5',       # 1ª - Azul (mais claro)
    'excel_red': 'C0504D',        # 2ª - Vermelho/Bordô
    'excel_green': '9BBB59',      # 3ª - Verde
    'excel_purple': '8064A2',     # 4ª - Roxo
    'excel_lightblue': '4BACC6',  # 5ª - Azul água
    'excel_orange': 'F79646'      # 6ª - Laranja
}

BORDER_ALL = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ═════════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES DE TRADUÇÃO
# ═════════════════════════════════════════════════════════════════

def translate_tipo(tipo, lang='pt'):
    """Traduzir tipo de atividade de grua"""
    if lang == 'en':
        translations = {
            'mobilizacao': 'Mobilization',
            'trabalho': 'Work',
            'paragem': 'Stoppage',
            'transferencia': 'Transfer',
            'desmobilizacao': 'Demobilization'
        }
    else:  # pt
        translations = {
            'mobilizacao': 'Mobilização',
            'trabalho': 'Trabalho',
            'paragem': 'Paragem',
            'transferencia': 'Transferência',
            'desmobilizacao': 'Desmobilização'
        }
    return translations.get(tipo, tipo)

def translate_motivo(motivo, lang='pt'):
    """Traduzir motivo de paragem"""
    if not motivo:
        return ''
    
    if lang == 'en':
        translations = {
            'wind': 'Wind',
            'mechanical': 'Mechanical Issue',
            'waiting_components': 'Waiting Components',
            'safety': 'Safety'
        }
    else:  # pt
        translations = {
            'wind': 'Vento',
            'mechanical': 'Problema Mecânico',
            'waiting_components': 'Aguardar Componentes',
            'safety': 'Segurança'
        }
    return translations.get(motivo, motivo)

def translate_phase(phase_key, lang='pt'):
    """Traduzir nome de fase"""
    if lang == 'en':
        translations = {
            'recepcao': 'Reception',
            'preparacao': 'Preparation',
            'preAssemblagem': 'Pre-Assembly',
            'assemblagem': 'Assembly',
            'torqueTensionamento': 'Torque & Tensioning',
            'fasesFinal': 'Final Phases',
            'ncrs': 'NCRs'
        }
    else:  # pt
        translations = {
            'recepcao': 'Receção',
            'preparacao': 'Preparação',
            'preAssemblagem': 'Pré-Assemblagem',
            'assemblagem': 'Assemblagem',
            'torqueTensionamento': 'Torque & Tensioning',
            'fasesFinal': 'Fases Finais',
            'ncrs': 'NCRs'
        }
    return translations.get(phase_key, phase_key)

def _add_ncr_sheet(wb, ncrs_data, language='pt'):
    ws = wb.create_sheet('NCRs')

    widths = [16, 24, 18, 16, 16, 16, 18, 14, 14, 16, 18, 14, 42, 36, 12, 28]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    headers = [
        'Código' if language == 'pt' else 'Code',
        'Título' if language == 'pt' else 'Title',
        'Turbina' if language == 'pt' else 'Turbine',
        'Categoria' if language == 'pt' else 'Category',
        'Severidade' if language == 'pt' else 'Severity',
        'Estado' if language == 'pt' else 'Status',
        'Responsável' if language == 'pt' else 'Assigned To',
        'Data Limite' if language == 'pt' else 'Due Date',
        'Criada' if language == 'pt' else 'Created',
        'Fechada' if language == 'pt' else 'Closed',
        'Fechada Por' if language == 'pt' else 'Closed By',
        'Evidências' if language == 'pt' else 'Evidence',
        'Descrição' if language == 'pt' else 'Description',
        'Nota de Fecho' if language == 'pt' else 'Closure Note',
        'Histórico' if language == 'pt' else 'History',
        'Última Nota de Estado' if language == 'pt' else 'Latest Status Note',
    ]

    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_ALL

    for row_idx, item in enumerate(ncrs_data, 2):
        ws.cell(row=row_idx, column=1, value=item.get('code', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=2, value=item.get('title', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=3, value=item.get('turbina', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=4, value=item.get('category', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=5, value=item.get('severity', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=6, value=item.get('status', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=7, value=item.get('assignedTo', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=8, value=item.get('dueDate', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=9, value=item.get('createdAt', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=10, value=item.get('closedAt', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=11, value=item.get('closedByName', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=12, value=item.get('evidenceCount', 0)).border = BORDER_ALL
        ws.cell(row=row_idx, column=13, value=item.get('description', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=14, value=item.get('closureNote', '')).border = BORDER_ALL
        ws.cell(row=row_idx, column=15, value=item.get('historyCount', 0)).border = BORDER_ALL
        ws.cell(row=row_idx, column=16, value=item.get('latestStatusNote', '')).border = BORDER_ALL

    ws.auto_filter.ref = f"A1:P{max(len(ncrs_data) + 1, 2)}"
    ws.freeze_panes = 'A2'
    print(f"✓ NCR sheet created with {len(ncrs_data)} items")

def get_status_icon(status):
    """Retornar ícone de status"""
    if status == 'completed' or status == 'concluído':
        return '✅'
    elif status == 'pending' or status == 'pendente':
        return '⏳'
    elif status == 'in_progress' or status == 'em_progresso':
        return '🔄'
    elif status == 'delayed' or status == 'atrasado':
        return '⚠️'
    else:
        return '⊘'

def parse_duration(duration_value):
    """Parse duration value, handling formats like '2h', '2.5', etc."""
    if not duration_value:
        return 0
    
    try:
        # Convert to string first
        duration_str = str(duration_value).strip().lower()
        # Remove 'h' or 'hora' suffixes
        duration_str = duration_str.replace('h', '').replace('hora', '').strip()
        # Convert to float
        return float(duration_str) if duration_str else 0
    except (ValueError, TypeError):
        return 0

def parse_percentage(percentage_value):
    """Parse percentage value, handling formats like '50%', '50', etc."""
    if percentage_value is None:
        return 0
    
    try:
        # Convert to string and remove % if present
        pct_str = str(percentage_value).strip().replace('%', '').strip()
        # Convert to int
        return int(float(pct_str)) if pct_str else 0
    except (ValueError, TypeError):
        return 0

# ═════════════════════════════════════════════════════════════════
# CRIAÇÃO DE COVER SHEET
# ═════════════════════════════════════════════════════════════════

def _create_cover_sheet(wb, project_name, language='pt'):
    """Criar sheet de capa profissional"""
    
    ws = wb.create_sheet('Capa', 0)
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 5
    
    # Título
    ws.merge_cells('B2:C2')
    title_cell = ws['B2']
    title_cell.value = '🏗️ AS-BUILT'
    title_cell.font = Font(name='Arial', size=20, bold=True, color=COLORS['header'])
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtítulo
    ws.merge_cells('B3:C3')
    subtitle_cell = ws['B3']
    subtitle_cell.value = 'Relatório de Instalação de Turbinas' if language == 'pt' else 'Wind Turbine Installation Report'
    subtitle_cell.font = Font(name='Arial', size=12, color=COLORS['dark_gray'])
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws['B4'].value = '─' * 50
    ws['B4'].alignment = Alignment(horizontal='center')
    
    current_row = 6
    
    # Nome do Projeto
    ws.merge_cells(f'B{current_row}:C{current_row}')
    project_cell = ws[f'B{current_row}']
    project_cell.value = project_name
    project_cell.font = Font(name='Arial', size=18, bold=True, color=COLORS['white'])
    project_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    project_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 35
    
    current_row += 2
    
    # Data do Relatório
    label_font = Font(name='Arial', size=10, bold=True, color=COLORS['dark_gray'])
    info_font = Font(name='Arial', size=11, color=COLORS['header'])
    
    ws[f'B{current_row}'].value = 'Data do Relatório:' if language == 'pt' else 'Report Date:'
    ws[f'B{current_row}'].font = label_font
    ws[f'C{current_row}'].value = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws[f'C{current_row}'].font = info_font
    
    print(f"✓ Cover sheet created")

# ═════════════════════════════════════════════════════════════════
# CRIAÇÃO DE EXECUTIVE SUMMARY
# ═════════════════════════════════════════════════════════════════

def _create_executive_summary_v2(wb, project_name, data_by_phase, language='pt'):
    """Criar sheet de resumo executivo com KPIs"""
    
    ws = wb.create_sheet('Resumo Executivo', 1)
    
    ws.column_dimensions['A'].width = 5
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Header
    ws.merge_cells('B1:G1')
    header = ws['B1']
    header.value = 'RESUMO EXECUTIVO' if language == 'pt' else 'EXECUTIVE SUMMARY'
    header.font = Font(name='Arial', size=14, bold=True, color=COLORS['white'])
    header.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header.alignment = Alignment(horizontal='center', vertical='center')
    
    # Calcular KPIs baseado em turbinas únicas
    turbine_ids = set()
    turbine_status = {}
    
    # Coletar turbinas únicas e seus status de todas as fases relevantes
    phase_keys = ['recepcao', 'preparacao', 'preAssemblagem', 'assemblagem', 'torqueTensionamento', 'fasesFinal']
    
    for phase_key in phase_keys:
        phase_data = data_by_phase.get(phase_key, [])
        if isinstance(phase_data, list):
            for item in phase_data:
                # Tentar obter ID da turbina (pode ser turbinaId, turbina, numero, etc.)
                turbine_id = item.get('turbinaId') or item.get('turbina') or item.get('numero') or item.get('id')
                if turbine_id:
                    turbine_ids.add(turbine_id)
                    # Atualizar status da turbina (último status encontrado)
                    status = item.get('status', '').lower().strip()
                    if status:
                        turbine_status[turbine_id] = status
    
    total_turbines = len(turbine_ids) if turbine_ids else 0
    
    # Contar turbinas concluídas
    completed_turbines = sum(1 for status in turbine_status.values() 
                             if status in ('concluído', 'concluido', 'completed', 'complete'))
    
    completion_rate = (completed_turbines / total_turbines * 100) if total_turbines > 0 else 0
    
    # KPIs Table
    row = 3
    
    # Cabeçalho da tabela de KPIs
    headers = ['Métrica', 'Valor', 'Meta', 'Status']
    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col+1)
        cell.value = header_text
        cell.font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_ALL
    
    row += 1
    
    # KPI 1: Total de Turbinas
    ws.cell(row=row, column=2).value = 'Total de Turbinas' if language == 'pt' else 'Total Turbines'
    ws.cell(row=row, column=3).value = total_turbines
    ws.cell(row=row, column=4).value = total_turbines
    status_cell = ws.cell(row=row, column=5)
    status_cell.value = '✅'
    ws.cell(row=row, column=2).border = BORDER_ALL
    ws.cell(row=row, column=3).border = BORDER_ALL
    ws.cell(row=row, column=4).border = BORDER_ALL
    ws.cell(row=row, column=5).border = BORDER_ALL
    row += 1
    
    # KPI 2: Taxa de Conclusão
    ws.cell(row=row, column=2).value = 'Taxa de Conclusão' if language == 'pt' else 'Completion Rate'
    ws.cell(row=row, column=3).value = f"{completion_rate:.1f}%"
    ws.cell(row=row, column=4).value = "100%"
    status_cell = ws.cell(row=row, column=5)
    if completion_rate >= 100:
        status_cell.value = '✅'
        status_cell.fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    elif completion_rate >= 80:
        status_cell.value = '🟡'
        status_cell.fill = PatternFill(start_color=COLORS['caution'], end_color=COLORS['caution'], fill_type='solid')
    else:
        status_cell.value = '⚠️'
        status_cell.fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
    
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = BORDER_ALL
    
    row += 1
    
    # KPI 3: Turbinas Concluídas
    ws.cell(row=row, column=2).value = 'Turbinas Concluídas' if language == 'pt' else 'Completed Turbines'
    ws.cell(row=row, column=3).value = f"{completed_turbines}/{total_turbines}"
    ws.cell(row=row, column=4).value = total_turbines
    status_cell = ws.cell(row=row, column=5)
    status_cell.value = '✅' if completed_turbines == total_turbines else '⏳'
    
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = BORDER_ALL
    
    print(f"✓ Executive summary created with KPIs")

# ═════════════════════════════════════════════════════════════════
# CRIAÇÃO DE DASHBOARD COM GRÁFICOS
# ═════════════════════════════════════════════════════════════════

def _create_dashboard_v3(wb, project_name, data_by_phase, language='pt'):
    """Dashboard com layout fixo: tabelas à esquerda (A-D), gráficos à direita (F+)"""
    
    ws = wb.create_sheet('Dashboard', 2)
    
    # Larguras das colunas - layout limpo: tabelas esquerda, gráficos direita
    ws.column_dimensions['A'].width = 22  # Nomes
    ws.column_dimensions['B'].width = 12  # Valores
    ws.column_dimensions['C'].width = 12  # Segundos valores
    ws.column_dimensions['D'].width = 12  # Percentagens/extras
    ws.column_dimensions['E'].width = 3   # Espaço separador
    for col in range(6, 16):  # F-O para gráficos
        ws.column_dimensions[get_column_letter(col)].width = 3
    
    # Header principal
    ws.merge_cells('A1:O1')
    header = ws['A1']
    header.value = f'DASHBOARD - {project_name}'
    header.font = Font(name='Arial', size=14, bold=True, color=COLORS['white'])
    header.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    def _normalize_status(value):
        return str(value or '').strip().lower()
    
    def _parse_date(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        text = str(value).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(text[:10], fmt)
            except Exception:
                continue
        return None
    
    # ═══════════════════════════════════════════════════════════════════════
    # COLETAR TODOS OS DADOS
    # ═══════════════════════════════════════════════════════════════════════
    
    # Fases - sempre 6 fases
    phase_keys = ['recepcao', 'preparacao', 'preAssemblagem', 'assemblagem', 'torqueTensionamento', 'fasesFinal']
    phase_rows = []
    
    # Coletar turbinas únicas e seus status
    turbine_ids = set()
    turbine_status = {}
    
    # Primeiro, coletar todas as turbinas únicas
    for phase_key in phase_keys:
        phase_data = data_by_phase.get(phase_key, [])
        if isinstance(phase_data, list):
            for item in phase_data:
                turbine_id = item.get('turbinaId') or item.get('turbina') or item.get('numero') or item.get('id')
                if turbine_id:
                    turbine_ids.add(turbine_id)
                    status = _normalize_status(item.get('status'))
                    if status:
                        turbine_status[turbine_id] = status
    
    total_turbines = len(turbine_ids) if turbine_ids else 0
    
    # Contar turbinas por status
    completed_turbines = sum(1 for status in turbine_status.values() 
                             if status in ('concluído', 'concluido', 'completed', 'complete'))
    in_progress_turbines = sum(1 for status in turbine_status.values() 
                               if status in ('em_progresso', 'em progresso', 'in_progress', 'in progress'))
    planned_turbines = total_turbines - completed_turbines - in_progress_turbines
    
    # Calcular dados por fase (número de componentes/itens em cada fase)
    for phase_key in phase_keys:
        phase_data = data_by_phase.get(phase_key, [])
        total = len(phase_data) if isinstance(phase_data, list) else 0
        concluded = 0
        if isinstance(phase_data, list):
            for item in phase_data:
                status = _normalize_status(item.get('status'))
                if status in ('concluído', 'concluido', 'completed', 'complete'):
                    concluded += 1
        pct = (concluded / total * 100) if total > 0 else 0
        phase_rows.append((translate_phase(phase_key, language), total, concluded, pct))
    
    completion_rate = (completed_turbines / total_turbines * 100) if total_turbines > 0 else 0
    
    # Timeline de instalação
    timeline_counts = defaultdict(int)
    date_fields = ['dataReal', 'dataAtual', 'dataFim', 'dataConclusao', 'dataInstalacao', 'data']
    for phase_data in data_by_phase.values():
        if isinstance(phase_data, list):
            for item in phase_data:
                date_value = None
                for field in date_fields:
                    date_value = _parse_date(item.get(field))
                    if date_value:
                        break
                if date_value:
                    year, week, _ = date_value.isocalendar()
                    label = f"Semana {week}" if language == 'pt' else f"Week {week}"
                    timeline_counts[label] += 1
    
    timeline_rows = sorted(timeline_counts.items(), key=lambda x: int(x[0].split()[-1]))
    if not timeline_rows:
        timeline_rows = [('Semana 1' if language == 'pt' else 'Week 1', 0)]
    
    # Gruas - trabalho vs paragens
    work_hours = 0.0
    stop_hours = 0.0
    for item in data_by_phase.get('gruasPads', []) + data_by_phase.get('gruasGerais', []):
        tipo = _normalize_status(item.get('tipo'))
        duracao = parse_duration(item.get('duracao'))
        if tipo in ('trabalho', 'work'):
            work_hours += duracao
        elif tipo in ('paragem', 'parada', 'stoppage', 'stop'):
            stop_hours += duracao
    
    # ═══════════════════════════════════════════════════════════════════════
    # SECÇÃO 1: PROGRESSO POR FASE (Linhas 3-10)
    # ═══════════════════════════════════════════════════════════════════════
    
    row = 3
    ws.merge_cells(f'A{row}:D{row}')
    title = ws[f'A{row}']
    title.value = 'PROGRESSO POR FASE' if language == 'pt' else 'PROGRESS BY PHASE'
    title.font = Font(name='Arial', size=11, bold=True, color=COLORS['header'])
    title.alignment = Alignment(horizontal='left', vertical='center')
    
    row = 4
    headers = ['Fase', 'Total', 'Concluídos', '%'] if language == 'pt' else ['Phase', 'Total', 'Completed', '%']
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = hdr
        cell.font = Font(bold=True, color=COLORS['white'], size=10)
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Cores para as fases - MESMA ORDEM que o Excel aplica nos gráficos
    phase_color_palette = [
        COLORS['excel_blue'],      # Receção - Azul
        COLORS['excel_red'],       # Preparação - Vermelho
        COLORS['excel_green'],     # Pré-Assemblagem - Verde
        COLORS['excel_purple'],    # Assemblagem - Roxo
        COLORS['excel_lightblue'], # Torque - Azul água
        COLORS['excel_orange']     # Fases Finais - Laranja
    ]
    
    for i, (phase_name, total, concluded, pct) in enumerate(phase_rows):
        row = 5 + i
        # Aplicar cor à célula do nome da fase (mesma cor que aparecerá no gráfico)
        color = phase_color_palette[i % len(phase_color_palette)]
        cell_name = ws.cell(row=row, column=1, value=phase_name)
        cell_name.border = BORDER_ALL
        cell_name.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell_name.font = Font(color=COLORS['white'], bold=True)
        
        ws.cell(row=row, column=2, value=total).border = BORDER_ALL
        ws.cell(row=row, column=3, value=concluded).border = BORDER_ALL
        ws.cell(row=row, column=4, value=f"{pct:.0f}%").border = BORDER_ALL
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    # Gráfico de barras (sem título)
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = None  # Remover título
    chart1.y_axis.title = ''
    chart1.x_axis.title = ''
    chart1.legend = None
    chart1.varyColors = True  # Usar cores diferentes para cada barra
    max_phase_row = 4 + len(phase_rows)
    data1 = Reference(ws, min_col=2, min_row=4, max_row=max_phase_row)
    cats1 = Reference(ws, min_col=1, min_row=5, max_row=max_phase_row)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.height = 8
    chart1.width = 14
    ws.add_chart(chart1, 'F3')
    
    # ═══════════════════════════════════════════════════════════════════════
    # SECÇÃO 2: STATUS DAS TURBINAS (Linhas 20-24)
    # ═══════════════════════════════════════════════════════════════════════
    
    row = 20
    ws.merge_cells(f'A{row}:D{row}')
    title = ws[f'A{row}']
    title.value = 'STATUS DAS TURBINAS' if language == 'pt' else 'TURBINE STATUS'
    title.font = Font(name='Arial', size=11, bold=True, color=COLORS['header'])
    title.alignment = Alignment(horizontal='left', vertical='center')
    
    row = 21
    status_headers = ['Status', 'Quantidade'] if language == 'pt' else ['Status', 'Count']
    for col, hdr in enumerate(status_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = hdr
        cell.font = Font(bold=True, color=COLORS['white'], size=10)
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    status_data = [
        ('Finalizadas' if language == 'pt' else 'Completed', completed_turbines),
        ('Em Instalação' if language == 'pt' else 'In Installation', in_progress_turbines),
        ('Planeadas' if language == 'pt' else 'Planned', planned_turbines)
    ]
    
    # Cores para os status - MESMA ORDEM que o Excel aplica no gráfico pizza
    status_colors = [
        COLORS['excel_blue'],    # Finalizadas - Azul (1ª cor)
        COLORS['excel_red'],     # Em Instalação - Vermelho (2ª cor)
        COLORS['excel_green']    # Planeadas - Verde (3ª cor)
    ]
    
    for i, (status_name, count) in enumerate(status_data):
        row = 22 + i
        # Aplicar cor à célula do status (mesma cor que aparecerá no gráfico)
        cell_status = ws.cell(row=row, column=1, value=status_name)
        cell_status.border = BORDER_ALL
        cell_status.fill = PatternFill(start_color=status_colors[i], end_color=status_colors[i], fill_type='solid')
        cell_status.font = Font(color=COLORS['white'], bold=True)
        cell_status.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=row, column=2, value=count).border = BORDER_ALL
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    # Gráfico de pizza 3D (sem título e sem legenda)
    chart2 = PieChart3D()
    chart2.title = None  # Remover título
    chart2.legend = None  # Remover legenda (as cores das células servem de guia)
    chart2.varyColors = True  # Usar cores diferentes para cada fatia
    max_status_row = 21 + len(status_data)
    data2 = Reference(ws, min_col=2, min_row=22, max_row=max_status_row)
    cats2 = Reference(ws, min_col=1, min_row=22, max_row=max_status_row)
    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(cats2)
    chart2.height = 8
    chart2.width = 14
    ws.add_chart(chart2, 'F20')
    
    # ═══════════════════════════════════════════════════════════════════════
    # SECÇÃO 3: TIMELINE DE INSTALAÇÃO (Linhas 37+)
    # ═══════════════════════════════════════════════════════════════════════
    
    row = 37
    ws.merge_cells(f'A{row}:D{row}')
    title = ws[f'A{row}']
    title.value = 'TIMELINE DE INSTALAÇÃO' if language == 'pt' else 'INSTALLATION TIMELINE'
    title.font = Font(name='Arial', size=11, bold=True, color=COLORS['header'])
    title.alignment = Alignment(horizontal='left', vertical='center')
    
    row = 38
    timeline_headers = ['Período', 'Turbinas'] if language == 'pt' else ['Period', 'Turbines']
    for col, hdr in enumerate(timeline_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = hdr
        cell.font = Font(bold=True, color=COLORS['white'], size=10)
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    timeline_start = 39
    # Cores para timeline - MESMA ORDEM que o Excel aplica no gráfico de linha
    timeline_color_palette = [
        COLORS['excel_blue'],      # Semana 1 - Azul
        COLORS['excel_red'],       # Semana 2 - Vermelho
        COLORS['excel_green'],     # Semana 3 - Verde
        COLORS['excel_purple'],    # Semana 4 - Roxo
        COLORS['excel_lightblue'], # Semana 5 - Azul água
        COLORS['excel_orange']     # Semana 6 - Laranja
    ]
    
    for i, (period, count) in enumerate(timeline_rows):
        row = timeline_start + i
        # Aplicar cor à célula do período (mesma cor que aparecerá no gráfico)
        color = timeline_color_palette[i % len(timeline_color_palette)]
        cell_period = ws.cell(row=row, column=1, value=period)
        cell_period.border = BORDER_ALL
        cell_period.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell_period.font = Font(color=COLORS['white'], bold=True)
        cell_period.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=row, column=2, value=count).border = BORDER_ALL
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    timeline_end = timeline_start + len(timeline_rows) - 1
    
    # Gráfico de linha (sem título e sem legenda)
    chart3 = LineChart()
    chart3.title = None  # Remover título
    chart3.legend = None  # Remover legenda (as cores das células servem de guia)
    chart3.y_axis.title = ''
    chart3.x_axis.title = ''
    chart3.varyColors = False  # Linha única, não variar cores
    data3 = Reference(ws, min_col=2, min_row=timeline_start, max_row=timeline_end)
    cats3 = Reference(ws, min_col=1, min_row=timeline_start, max_row=timeline_end)
    chart3.add_data(data3, titles_from_data=False)
    chart3.set_categories(cats3)
    chart3.height = 7
    chart3.width = 14
    ws.add_chart(chart3, 'F37')
    
    # ═══════════════════════════════════════════════════════════════════════
    # SECÇÃO 4: GRUAS - TRABALHO vs PARAGENS (Linha 54)
    # ═══════════════════════════════════════════════════════════════════════
    
    row = 54
    ws.merge_cells(f'A{row}:D{row}')
    title = ws[f'A{row}']
    title.value = 'GRUAS - TRABALHO vs PARAGENS' if language == 'pt' else 'CRANES - WORK vs STOPPAGES'
    title.font = Font(name='Arial', size=11, bold=True, color=COLORS['header'])
    title.alignment = Alignment(horizontal='left', vertical='center')
    
    row = 55
    gruas_headers = ['Tipo', 'Horas'] if language == 'pt' else ['Type', 'Hours']
    for col, hdr in enumerate(gruas_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = hdr
        cell.font = Font(bold=True, color=COLORS['white'], size=10)
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    gruas_data = [
        ('Trabalho' if language == 'pt' else 'Work', round(work_hours, 2)),
        ('Paragens' if language == 'pt' else 'Stoppages', round(stop_hours, 2))
    ]
    
    # Cores para gruas - MESMA ORDEM que o Excel aplica no gráfico
    gruas_colors = [
        COLORS['excel_blue'],    # Trabalho - Azul (1ª cor)
        COLORS['excel_red']      # Paragens - Vermelho (2ª cor)
    ]
    
    for i, (label, hours) in enumerate(gruas_data):
        row = 56 + i
        # Aplicar cor à célula do tipo (mesma cor que aparecerá no gráfico)
        cell_tipo = ws.cell(row=row, column=1, value=label)
        cell_tipo.border = BORDER_ALL
        cell_tipo.fill = PatternFill(start_color=gruas_colors[i], end_color=gruas_colors[i], fill_type='solid')
        cell_tipo.font = Font(color=COLORS['white'], bold=True)
        cell_tipo.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=row, column=2, value=hours).border = BORDER_ALL
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    # Gráfico de barras (sem título)
    chart4 = BarChart()
    chart4.type = "col"
    chart4.style = 10
    chart4.title = None  # Remover título
    chart4.y_axis.title = ''
    chart4.x_axis.title = ''
    chart4.legend = None
    chart4.varyColors = True  # Usar cores diferentes para cada barra
    max_gruas_row = 55 + len(gruas_data)
    data4 = Reference(ws, min_col=2, min_row=56, max_row=max_gruas_row)
    cats4 = Reference(ws, min_col=1, min_row=56, max_row=max_gruas_row)
    chart4.add_data(data4, titles_from_data=False)
    chart4.set_categories(cats4)
    chart4.height = 7
    chart4.width = 14
    ws.add_chart(chart4, 'F54')
    
    # ═══════════════════════════════════════════════════════════════════════
    # SECÇÃO 5: KPIs PRINCIPAIS (Linha 71)
    # ═══════════════════════════════════════════════════════════════════════
    
    row = 71
    ws.merge_cells(f'A{row}:D{row}')
    title = ws[f'A{row}']
    title.value = 'KPIs PRINCIPAIS' if language == 'pt' else 'MAIN KPIs'
    title.font = Font(name='Arial', size=11, bold=True, color=COLORS['header'])
    title.alignment = Alignment(horizontal='left', vertical='center')
    
    row = 72
    kpi_headers = ['KPI', 'Valor', 'Meta', 'Status'] if language == 'pt' else ['KPI', 'Value', 'Target', 'Status']
    for col, hdr in enumerate(kpi_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = hdr
        cell.font = Font(bold=True, color=COLORS['white'], size=10)
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    efficiency = (work_hours / (work_hours + stop_hours) * 100) if (work_hours + stop_hours) > 0 else 0
    kpis = [
        ('Total Concluído' if language == 'pt' else 'Completed Total', completion_rate, 100),
        ('Eficiência Gruas' if language == 'pt' else 'Crane Efficiency', efficiency, 85),
        ('Progresso Geral' if language == 'pt' else 'Overall Progress', completion_rate, 100)
    ]
    
    # Cores para KPIs - MESMA ORDEM que o Excel aplicaria
    kpi_color_palette = [
        COLORS['excel_blue'],      # Total Concluído - Azul
        COLORS['excel_red'],       # Eficiência Gruas - Vermelho
        COLORS['excel_green']      # Progresso Geral - Verde
    ]
    
    for i, (kpi_name, value, target) in enumerate(kpis):
        row = 73 + i
        # Aplicar cor à célula do nome do KPI (mesma cor do padrão Excel)
        color = kpi_color_palette[i % len(kpi_color_palette)]
        cell_kpi = ws.cell(row=row, column=1, value=kpi_name)
        cell_kpi.border = BORDER_ALL
        cell_kpi.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell_kpi.font = Font(color=COLORS['white'], bold=True)
        cell_kpi.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=row, column=2, value=f"{value:.0f}%").border = BORDER_ALL
        ws.cell(row=row, column=3, value=f"{target}%").border = BORDER_ALL
        status_cell = ws.cell(row=row, column=4)
        status_cell.value = '✅' if value >= target else '❌'
        status_cell.border = BORDER_ALL
        for col in range(2, 5):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"✓ Dashboard: {completed_turbines}/{total_turbines} turbinas ({completion_rate:.0f}%), {len(timeline_rows)} períodos")

# ═════════════════════════════════════════════════════════════════
# CRIAÇÃO DE ANÁLISE DE DESVIOS
# ═════════════════════════════════════════════════════════════════

def _add_deviation_analysis_sheet(wb, data_by_phase, language='pt'):
    """Criar sheet de análise de desvios (Planned vs Actual)"""
    
    ws = wb.create_sheet('Análise de Desvios' if language == 'pt' else 'Deviation Analysis')
    
    # Configurar larguras
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22
    
    # Header
    ws.merge_cells('A1:F1')
    header = ws['A1']
    header.value = '📈 ANÁLISE DE DESVIOS' if language == 'pt' else '📈 DEVIATION ANALYSIS'
    header.font = Font(name='Arial', size=12, bold=True, color=COLORS['white'])
    header.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header.alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalhos das colunas
    headers = ['Turbina', 'Fase', 'Planeado', 'Real', 'Desvio (dias)', 'Status']
    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header_text
        cell.font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_ALL
    
    # Coletar dados de desvios
    row = 3
    deviations = []
    
    for phase_key, phase_data in data_by_phase.items():
        if phase_data and isinstance(phase_data, list):
            for item in phase_data:
                turbine_name = item.get('turbina', item.get('numero', ''))
                planned_date = item.get('dataPlaneada', item.get('dataPrevista', ''))
                actual_date = item.get('dataReal', item.get('dataAtual', ''))
                
                if planned_date and actual_date:
                    try:
                        # Converter para datetime (simplificado)
                        planned = datetime.strptime(str(planned_date)[:10], '%Y-%m-%d') if isinstance(planned_date, str) else planned_date
                        actual = datetime.strptime(str(actual_date)[:10], '%Y-%m-%d') if isinstance(actual_date, str) else actual_date
                        
                        deviation_days = (actual - planned).days
                        status = 'No plan' if deviation_days == 0 else ('Adiantado' if deviation_days < 0 else 'Atrasado')
                        
                        deviations.append({
                            'turbine': turbine_name,
                            'phase': translate_phase(phase_key, language),
                            'planned': planned_date,
                            'actual': actual_date,
                            'deviation': deviation_days,
                            'status': status
                        })
                    except:
                        pass
    
    # Escrever dados
    for dev in sorted(deviations, key=lambda x: x['turbine']):
        ws.cell(row=row, column=1).value = dev['turbine']
        ws.cell(row=row, column=2).value = dev['phase']
        ws.cell(row=row, column=3).value = str(dev['planned'])
        ws.cell(row=row, column=4).value = str(dev['actual'])
        
        deviation_cell = ws.cell(row=row, column=5)
        deviation_cell.value = dev['deviation']
        deviation_cell.alignment = Alignment(horizontal='center')
        
        # Conditional formatting para desvios
        if dev['deviation'] > 0:
            deviation_cell.fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
        elif dev['deviation'] < 0:
            deviation_cell.fill = PatternFill(start_color=COLORS['caution'], end_color=COLORS['caution'], fill_type='solid')
        
        status_cell = ws.cell(row=row, column=6)
        status_cell.value = dev['status']
        status_cell.alignment = Alignment(horizontal='center')
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER_ALL
        
        row += 1
    
    # Auto-filtro
    if row > 3:
        ws.auto_filter.ref = f"A2:F{row-1}"
    
    # Freeze panes
    ws.freeze_panes = 'A3'
    
    print(f"✓ Deviation analysis sheet created with {len(deviations)} deviations")

# ═════════════════════════════════════════════════════════════════
# CRIAÇÃO DE ANÁLISE DE GRUAS
# ═════════════════════════════════════════════════════════════════

def _add_crane_analysis_sheet(wb, gruas_pads_data, gruas_gerais_data, language='pt'):
    """Criar sheet de análise detalhada de utilização de gruas"""
    
    ws = wb.create_sheet('Gruas - Análise' if language == 'pt' else 'Cranes - Analysis')
    
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Header
    ws.merge_cells('A1:I1')
    header = ws['A1']
    header.value = '🏗️ ANÁLISE DE UTILIZAÇÃO DE GRUAS' if language == 'pt' else '🏗️ CRANE UTILIZATION ANALYSIS'
    header.font = Font(name='Arial', size=12, bold=True, color=COLORS['white'])
    header.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header.alignment = Alignment(horizontal='center', vertical='center')
    
    # Section 1: Resumo de Gruas
    row = 3
    ws.merge_cells(f'A{row}:I{row}')
    section_header = ws[f'A{row}']
    section_header.value = 'Resumo de Utilização por Grua' if language == 'pt' else 'Crane Utilization Summary'
    section_header.font = Font(name='Arial', size=11, bold=True, color=COLORS['white'])
    section_header.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    section_header.alignment = Alignment(horizontal='center')
    
    row += 1
    
    # Cabeçalhos
    headers = ['Grua', 'Trabalho (h)', 'Mobiliz. (h)', 'Paragens (h)', 'Total (h)', 'Eficiência %', 'Status', 'Observações', 'META']
    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header_text
        cell.font = Font(name='Arial', size=9, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_ALL
    
    # Agregar dados de gruas
    crane_stats = defaultdict(lambda: {'trabalho': 0, 'mobilizacao': 0, 'paragem': 0})
    
    for grua_list in [gruas_pads_data, gruas_gerais_data]:
        if grua_list:
            for grua in grua_list:
                modelo = grua.get('gruaModelo', 'Desconhecida')
                tipo = grua.get('tipo', '')
                duracao = parse_duration(grua.get('duracao'))
                
                if tipo == 'trabalho':
                    crane_stats[modelo]['trabalho'] += duracao
                elif tipo == 'mobilizacao':
                    crane_stats[modelo]['mobilizacao'] += duracao
                elif tipo == 'paragem':
                    crane_stats[modelo]['paragem'] += duracao
    
    row += 1
    
    # Escrever dados agregados
    for crane_name, stats in sorted(crane_stats.items()):
        trabalho = stats['trabalho']
        mobilizacao = stats['mobilizacao']
        paragem = stats['paragem']
        total = trabalho + mobilizacao + paragem
        
        eficiencia = (trabalho / total * 100) if total > 0 else 0
        meta_eficiencia = 85
        
        ws.cell(row=row, column=1).value = crane_name
        ws.cell(row=row, column=2).value = round(trabalho, 2)
        ws.cell(row=row, column=3).value = round(mobilizacao, 2)
        ws.cell(row=row, column=4).value = round(paragem, 2)
        ws.cell(row=row, column=5).value = round(total, 2)
        
        efic_cell = ws.cell(row=row, column=6)
        efic_cell.value = f"{eficiencia:.1f}%"
        efic_cell.alignment = Alignment(horizontal='center')
        
        # Conditional formatting para eficiência
        if eficiencia >= meta_eficiencia:
            efic_cell.fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
        elif eficiencia >= 70:
            efic_cell.fill = PatternFill(start_color=COLORS['caution'], end_color=COLORS['caution'], fill_type='solid')
        else:
            efic_cell.fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
        
        status_cell = ws.cell(row=row, column=7)
        status_cell.value = '✅' if eficiencia >= meta_eficiencia else ('⚠️' if eficiencia >= 70 else '❌')
        status_cell.alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=8).value = 'OK' if eficiencia >= meta_eficiencia else 'Revisar'
        ws.cell(row=row, column=9).value = f"{meta_eficiencia}%"
        
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = BORDER_ALL
        
        row += 1
    
    # Section 2: Motivos de Paragem
    row += 2
    ws.merge_cells(f'A{row}:I{row}')
    section_header = ws[f'A{row}']
    section_header.value = 'Análise de Motivos de Paragem' if language == 'pt' else 'Stoppage Reasons Analysis'
    section_header.font = Font(name='Arial', size=11, bold=True, color=COLORS['white'])
    section_header.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    
    row += 1
    
    # Cabeçalhos
    motivo_headers = ['Motivo', 'Frequência', 'Tempo Total (h)', 'Percentual']
    for col, header_text in enumerate(motivo_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header_text
        cell.font = Font(name='Arial', size=9, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER_ALL
    
    # Agregar motivos
    motivo_stats = defaultdict(lambda: {'count': 0, 'duracao': 0})
    
    for grua_list in [gruas_pads_data, gruas_gerais_data]:
        if grua_list:
            for grua in grua_list:
                if grua.get('tipo') == 'paragem':
                    motivo = grua.get('motivo', 'Não especificado')
                    duracao = parse_duration(grua.get('duracao'))
                    motivo_stats[motivo]['count'] += 1
                    motivo_stats[motivo]['duracao'] += duracao
    
    row += 1
    
    total_paragem_time = sum(v['duracao'] for v in motivo_stats.values())
    
    for motivo, stats in sorted(motivo_stats.items(), key=lambda x: x[1]['duracao'], reverse=True):
        ws.cell(row=row, column=1).value = translate_motivo(motivo, language)
        ws.cell(row=row, column=2).value = stats['count']
        ws.cell(row=row, column=3).value = round(stats['duracao'], 2)
        
        pct_cell = ws.cell(row=row, column=4)
        pct = (stats['duracao'] / total_paragem_time * 100) if total_paragem_time > 0 else 0
        pct_cell.value = f"{pct:.1f}%"
        pct_cell.alignment = Alignment(horizontal='center')
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BORDER_ALL
        
        row += 1
    
    # Auto-filtro na primeira seção
    ws.auto_filter.ref = f"A4:I{row-len(motivo_stats)-3}"
    
    # Freeze panes
    ws.freeze_panes = 'A3'
    
    print(f"✓ Crane analysis sheet created")

# ═════════════════════════════════════════════════════════════════
# CRIAÇÃO DE OBSERVAÇÕES CRÍTICAS
# ═════════════════════════════════════════════════════════════════

def _add_critical_observations_sheet(wb, data_by_phase, language='pt'):
    """Criar sheet de observações críticas e problemas"""
    
    ws = wb.create_sheet('Observações Críticas' if language == 'pt' else 'Critical Observations')
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Header
    ws.merge_cells('A1:F1')
    header = ws['A1']
    header.value = '⚠️ OBSERVAÇÕES CRÍTICAS' if language == 'pt' else '⚠️ CRITICAL OBSERVATIONS'
    header.font = Font(name='Arial', size=12, bold=True, color=COLORS['white'])
    header.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header.alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalhos
    headers = ['Data', 'Turbina', 'Fase', 'Categoria', 'Descrição', 'Status']
    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header_text
        cell.font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_ALL
    
    # Coletar observações
    observations = []
    
    for phase_key, phase_data in data_by_phase.items():
        if phase_data and isinstance(phase_data, list):
            for item in phase_data:
                obs = item.get('observacoes', '')
                if obs:
                    observations.append({
                        'data': item.get('data', datetime.now().strftime('%Y-%m-%d')),
                        'turbine': item.get('turbina', item.get('numero', '')),
                        'phase': translate_phase(phase_key, language),
                        'category': '⚠️ Crítico' if 'crítico' in str(obs).lower() else ('🔴 Problema' if 'problema' in str(obs).lower() else ('🟡 Atenção' if 'atenção' in str(obs).lower() else 'ℹ️ Info')),
                        'description': obs,
                        'status': 'Aberto' if 'aberto' in str(obs).lower() else 'Resolvido'
                    })
    
    # Escrever dados
    row = 3
    for obs in sorted(observations, key=lambda x: x['data'], reverse=True):
        ws.cell(row=row, column=1).value = obs['data']
        ws.cell(row=row, column=2).value = obs['turbine']
        ws.cell(row=row, column=3).value = obs['phase']
        ws.cell(row=row, column=4).value = obs['category']
        ws.cell(row=row, column=5).value = obs['description']
        
        status_cell = ws.cell(row=row, column=6)
        status_cell.value = obs['status']
        status_cell.alignment = Alignment(horizontal='center')
        
        # Colorir conforme status
        if 'Crítico' in obs['category']:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
        elif 'Problema' in obs['category']:
            ws.cell(row=row, column=4).fill = PatternFill(start_color=COLORS['caution'], end_color=COLORS['caution'], fill_type='solid')
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER_ALL
        
        row += 1
    
    # Auto-filtro
    if row > 3:
        ws.auto_filter.ref = f"A2:F{row-1}"
    
    # Freeze panes
    ws.freeze_panes = 'A3'
    
    print(f"✓ Critical observations sheet created with {len(observations)} observations")

# ═════════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES PARA FASES
# ═════════════════════════════════════════════════════════════════

def _add_project_header(ws, project_name, sheet_name, language='pt'):
    """Adicionar cabeçalho de projeto a uma sheet"""
    # Insert rows at top
    ws.insert_rows(1, 3)
    
    # Configurar altura das linhas de header
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    
    # Título
    ws.merge_cells('A1:Z1')
    title = ws['A1']
    title.value = project_name
    title.font = Font(name='Arial', size=12, bold=True, color=COLORS['white'])
    title.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Nome da sheet
    ws.merge_cells('A2:Z2')
    sheet_title = ws['A2']
    sheet_title.value = sheet_name
    sheet_title.font = Font(name='Arial', size=11, bold=True, color=COLORS['white'])
    sheet_title.fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    sheet_title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _add_footer(ws, language='pt'):
    """Adicionar rodapé a uma sheet"""
    last_row = ws.max_row + 2
    ws.merge_cells(f'A{last_row}:Z{last_row}')
    footer = ws[f'A{last_row}']
    footer.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    footer.font = Font(name='Arial', size=8, italic=True, color=COLORS['dark_gray'])
    footer.alignment = Alignment(horizontal='right', vertical='center')
    footer.fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')

def _add_auto_filters(ws, header_row, num_cols):
    """Adicionar auto-filtros a uma sheet"""
    start_col = get_column_letter(1)
    end_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"{start_col}{header_row}:{end_col}{ws.max_row}"
    ws.freeze_panes = f"A{header_row + 1}"

# ═════════════════════════════════════════════════════════════════
# FUNÇÕES PARA ADICIONAR SHEETS DE FASES
# ═════════════════════════════════════════════════════════════════

def _add_recepcao_sheet(wb, recepcao_data, lang='pt'):
    """Adicionar sheet de Receção"""
    ws = wb.create_sheet('Receção' if lang == 'pt' else 'Reception')
    
    border = BORDER_ALL
    header_fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
    
    # Cabeçalhos
    headers = ['Turbina', 'Componente', 'VUI', 'Nº Série', 'Item Number', 'Data Descarga', 'Hora Descarga']
    widths = [15, 20, 18, 18, 18, 15, 15]
    
    for col, (header_text, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Dados
    for row_idx, item in enumerate(recepcao_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('turbinaId', '')).border = border
        ws.cell(row=row_idx, column=2, value=item.get('componentId', '')).border = border
        ws.cell(row=row_idx, column=3, value=item.get('vui', '')).border = border
        ws.cell(row=row_idx, column=4, value=item.get('serialNumber', '')).border = border
        ws.cell(row=row_idx, column=5, value=item.get('itemNumber', '')).border = border
        ws.cell(row=row_idx, column=6, value=item.get('dataDescarga', '')).border = border
        ws.cell(row=row_idx, column=7, value=item.get('horaDescarga', '')).border = border
        
        # Alinhamento
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"✓ Reception sheet created with {len(recepcao_data)} items")

def _add_preparacao_sheet(wb, preparacao_data, lang='pt'):
    """Adicionar sheet de Preparação"""
    ws = wb.create_sheet('Preparação' if lang == 'pt' else 'Preparation')
    
    border = BORDER_ALL
    header_fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
    
    headers = ['Turbina', 'Componente', 'VUI', 'Nº Série', 'Item Number', 'Data Início', 'Hora Início', 'Data Fim', 'Hora Fim']
    widths = [15, 20, 18, 18, 18, 15, 12, 15, 12]
    
    for col, (header_text, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = width
    
    for row_idx, item in enumerate(preparacao_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('turbinaId', '')).border = border
        ws.cell(row=row_idx, column=2, value=item.get('componentId', '')).border = border
        ws.cell(row=row_idx, column=3, value=item.get('vui', '')).border = border
        ws.cell(row=row_idx, column=4, value=item.get('serialNumber', '')).border = border
        ws.cell(row=row_idx, column=5, value=item.get('itemNumber', '')).border = border
        ws.cell(row=row_idx, column=6, value=item.get('dataInicio', '')).border = border
        ws.cell(row=row_idx, column=7, value=item.get('horaInicio', '')).border = border
        ws.cell(row=row_idx, column=8, value=item.get('dataFim', '')).border = border
        ws.cell(row=row_idx, column=9, value=item.get('horaFim', '')).border = border
        
        # Alinhamento
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"✓ Preparation sheet created with {len(preparacao_data)} items")

def _add_pre_assemblagem_sheet(wb, pre_data, lang='pt'):
    """Adicionar sheet de Pré-Assemblagem"""
    ws = wb.create_sheet('Pré-Assemblagem' if lang == 'pt' else 'Pre-Assembly')
    
    border = BORDER_ALL
    header_fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
    
    headers = ['Turbina', 'Componente', 'VUI', 'Nº Série', 'Item Number', 'Data Início', 'Hora Início', 'Data Fim', 'Hora Fim']
    widths = [15, 20, 18, 18, 18, 15, 12, 15, 12]
    
    for col, (header_text, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = width
    
    for row_idx, item in enumerate(pre_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('turbinaId', '')).border = border
        ws.cell(row=row_idx, column=2, value=item.get('componentId', '')).border = border
        ws.cell(row=row_idx, column=3, value=item.get('vui', '')).border = border
        ws.cell(row=row_idx, column=4, value=item.get('serialNumber', '')).border = border
        ws.cell(row=row_idx, column=5, value=item.get('itemNumber', '')).border = border
        ws.cell(row=row_idx, column=6, value=item.get('dataInicio', '')).border = border
        ws.cell(row=row_idx, column=7, value=item.get('horaInicio', '')).border = border
        ws.cell(row=row_idx, column=8, value=item.get('dataFim', '')).border = border
        ws.cell(row=row_idx, column=9, value=item.get('horaFim', '')).border = border
        
        # Alinhamento
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"✓ Pre-Assembly sheet created with {len(pre_data)} items")

def _add_assemblagem_sheet(wb, assemblagem_data, lang='pt'):
    """Adicionar sheet de Assemblagem"""
    ws = wb.create_sheet('Assemblagem' if lang == 'pt' else 'Assembly')
    
    border = BORDER_ALL
    header_fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
    
    headers = ['Turbina', 'Componente', 'VUI', 'Nº Série', 'Item Number', 'Data Início', 'Hora Início', 'Data Fim', 'Hora Fim']
    widths = [15, 20, 18, 18, 18, 15, 12, 15, 12]
    
    for col, (header_text, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = width
    
    for row_idx, item in enumerate(assemblagem_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('turbinaId', '')).border = border
        ws.cell(row=row_idx, column=2, value=item.get('componentId', '')).border = border
        ws.cell(row=row_idx, column=3, value=item.get('vui', '')).border = border
        ws.cell(row=row_idx, column=4, value=item.get('serialNumber', '')).border = border
        ws.cell(row=row_idx, column=5, value=item.get('itemNumber', '')).border = border
        ws.cell(row=row_idx, column=6, value=item.get('dataInicio', '')).border = border
        ws.cell(row=row_idx, column=7, value=item.get('horaInicio', '')).border = border
        ws.cell(row=row_idx, column=8, value=item.get('dataFim', '')).border = border
        ws.cell(row=row_idx, column=9, value=item.get('horaFim', '')).border = border
        
        # Alinhamento
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"✓ Assembly sheet created with {len(assemblagem_data)} items")

def _add_torque_sheet(wb, torque_data, lang='pt'):
    """Adicionar sheet de Torque & Tensioning"""
    ws = wb.create_sheet('Torque & Tensionamento' if lang == 'pt' else 'Torque & Tensioning')
    
    border = BORDER_ALL
    header_fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
    
    headers = ['Número', 'Componente', 'Parafusos', 'Torque (Nm)', 'Status', 'Inspetor', 'Data', 'Observações']
    widths = [12, 15, 12, 15, 12, 15, 12, 25]
    
    for col, (header_text, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    for row_idx, item in enumerate(torque_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('turbina', '')).border = border
        ws.cell(row=row_idx, column=2, value=item.get('componente', '')).border = border
        ws.cell(row=row_idx, column=3, value=item.get('parafusos', '')).border = border
        ws.cell(row=row_idx, column=4, value=item.get('torque', '')).border = border
        ws.cell(row=row_idx, column=5, value=item.get('status', '')).border = border
        ws.cell(row=row_idx, column=6, value=item.get('inspetor', '')).border = border
        ws.cell(row=row_idx, column=7, value=item.get('data', '')).border = border
        ws.cell(row=row_idx, column=8, value=item.get('observacoes', '')).border = border
    
    print(f"✓ Torque sheet created with {len(torque_data)} items")

def _add_fases_finais_sheet(wb, fases_data, lang='pt'):
    """Adicionar sheet de Fases Finais"""
    ws = wb.create_sheet('Fases Finais' if lang == 'pt' else 'Final Phases')
    
    border = BORDER_ALL
    header_fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color=COLORS['white'])
    
    headers = ['Turbina', 'Componente', 'VUI', 'Nº Série', 'Item Number', 'Data Início', 'Hora Início', 'Data Fim', 'Hora Fim']
    widths = [15, 20, 18, 18, 18, 15, 12, 15, 12]
    
    for col, (header_text, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = width
    
    for row_idx, item in enumerate(fases_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('turbinaId', '')).border = border
        ws.cell(row=row_idx, column=2, value=item.get('componentId', '')).border = border
        ws.cell(row=row_idx, column=3, value=item.get('vui', '')).border = border
        ws.cell(row=row_idx, column=4, value=item.get('serialNumber', '')).border = border
        ws.cell(row=row_idx, column=5, value=item.get('itemNumber', '')).border = border
        ws.cell(row=row_idx, column=6, value=item.get('dataInicio', '')).border = border
        ws.cell(row=row_idx, column=7, value=item.get('horaInicio', '')).border = border
        ws.cell(row=row_idx, column=8, value=item.get('dataFim', '')).border = border
        ws.cell(row=row_idx, column=9, value=item.get('horaFim', '')).border = border
        
        # Alinhamento
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"✓ Final phases sheet created with {len(fases_data)} items")

def _add_gruas_pads_sheet(wb, gruas_data, lang='pt'):
    """Adicionar sheet de Gruas - Pads"""
    ws = wb.create_sheet('Gruas - Pads')
    
    border = BORDER_ALL
    header_fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    header_font = Font(name='Arial', size=9, bold=True, color=COLORS['white'])
    
    headers = ['Modelo', 'Descrição', 'Tipo', 'Data Início', 'Hora Início', 'Data Fim', 'Hora Fim', 'Duração (h)', 'Motivo', 'Origem', 'Destino', 'Observações']
    widths = [14, 16, 12, 14, 12, 14, 12, 12, 14, 14, 14, 20]
    
    for col, (header_text, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    
    for row_idx, item in enumerate(gruas_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('gruaModelo', '')).border = border
        ws.cell(row=row_idx, column=2, value=item.get('descricao', '')).border = border
        ws.cell(row=row_idx, column=3, value=translate_tipo(item.get('tipo', ''), lang)).border = border
        ws.cell(row=row_idx, column=4, value=item.get('dataInicio', '')).border = border
        ws.cell(row=row_idx, column=5, value=item.get('horaInicio', '')).border = border
        ws.cell(row=row_idx, column=6, value=item.get('dataFim', '')).border = border
        ws.cell(row=row_idx, column=7, value=item.get('horaFim', '')).border = border
        ws.cell(row=row_idx, column=8, value=item.get('duracao', '')).border = border
        ws.cell(row=row_idx, column=9, value=translate_motivo(item.get('motivo', ''), lang)).border = border
        ws.cell(row=row_idx, column=10, value=item.get('origem', '')).border = border
        ws.cell(row=row_idx, column=11, value=item.get('destino', '')).border = border
        ws.cell(row=row_idx, column=12, value=item.get('observacoes', '')).border = border
    
    print(f"✓ Gruas - Pads sheet created with {len(gruas_data)} items")

def _add_gruas_gerais_sheet(wb, gruas_data, lang='pt'):
    """Adicionar sheet de Gruas - Gerais"""
    ws = wb.create_sheet('Gruas - Gerais')
    
    border = BORDER_ALL
    header_fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    header_font = Font(name='Arial', size=9, bold=True, color=COLORS['white'])
    
    headers = ['Modelo', 'Descrição', 'Tipo', 'Data Início', 'Hora Início', 'Data Fim', 'Hora Fim', 'Duração (h)', 'Motivo', 'Origem', 'Destino', 'Observações']
    widths = [14, 16, 12, 14, 12, 14, 12, 12, 14, 14, 14, 20]
    
    for col, (header_text, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    
    for row_idx, item in enumerate(gruas_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('gruaModelo', '')).border = border
        ws.cell(row=row_idx, column=2, value=item.get('descricao', '')).border = border
        ws.cell(row=row_idx, column=3, value=translate_tipo(item.get('tipo', ''), lang)).border = border
        ws.cell(row=row_idx, column=4, value=item.get('dataInicio', '')).border = border
        ws.cell(row=row_idx, column=5, value=item.get('horaInicio', '')).border = border
        ws.cell(row=row_idx, column=6, value=item.get('dataFim', '')).border = border
        ws.cell(row=row_idx, column=7, value=item.get('horaFim', '')).border = border
        ws.cell(row=row_idx, column=8, value=item.get('duracao', '')).border = border
        ws.cell(row=row_idx, column=9, value=translate_motivo(item.get('motivo', ''), lang)).border = border
        ws.cell(row=row_idx, column=10, value=item.get('origem', '')).border = border
        ws.cell(row=row_idx, column=11, value=item.get('destino', '')).border = border
        ws.cell(row=row_idx, column=12, value=item.get('observacoes', '')).border = border
    
    print(f"✓ Gruas - Gerais sheet created with {len(gruas_data)} items")

def _add_equipamentos_sheet(wb, equipamentos_data, lang='pt'):
    """Adicionar sheet de Equipamentos (rastreabilidade)"""
    ws = wb.create_sheet('Equipamentos' if lang == 'pt' else 'Equipment')

    border = BORDER_ALL
    header_fill = PatternFill(start_color=COLORS['subheader'], end_color=COLORS['subheader'], fill_type='solid')
    header_font = Font(name='Arial', size=9, bold=True, color=COLORS['white'])

    headers = [
        'ID',
        'Tipo',
        'Fabricante' if lang == 'pt' else 'Manufacturer',
        'Modelo' if lang == 'pt' else 'Model',
        'Nº Série' if lang == 'pt' else 'Serial Number',
        'Estado' if lang == 'pt' else 'Status',
        'Condição' if lang == 'pt' else 'Condition',
        'Calibração' if lang == 'pt' else 'Calibration',
        'Validade' if lang == 'pt' else 'Expiry',
        'Certificado' if lang == 'pt' else 'Certificate',
        'Localização' if lang == 'pt' else 'Location',
        'Projeto Uso' if lang == 'pt' else 'Usage Project',
        'Turbina' if lang == 'pt' else 'Turbine',
        'Ligação' if lang == 'pt' else 'Connection',
        'Operação' if lang == 'pt' else 'Operation',
        'Utilizador' if lang == 'pt' else 'User',
        'Data Uso' if lang == 'pt' else 'Usage Date',
        'Observações' if lang == 'pt' else 'Notes',
    ]

    widths = [14, 16, 16, 16, 16, 12, 14, 12, 12, 14, 14, 16, 12, 14, 14, 14, 12, 20]

    for col, (header_text, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx, item in enumerate(equipamentos_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('equipmentId', '')).border = border
        ws.cell(row=row_idx, column=2, value=item.get('type', '')).border = border
        ws.cell(row=row_idx, column=3, value=item.get('manufacturer', '')).border = border
        ws.cell(row=row_idx, column=4, value=item.get('model', '')).border = border
        ws.cell(row=row_idx, column=5, value=item.get('serialNumber', '')).border = border
        ws.cell(row=row_idx, column=6, value=item.get('status', '')).border = border
        ws.cell(row=row_idx, column=7, value=item.get('condition', '')).border = border
        ws.cell(row=row_idx, column=8, value=item.get('calibrationLastDate', '')).border = border
        ws.cell(row=row_idx, column=9, value=item.get('calibrationExpiryDate', '')).border = border
        ws.cell(row=row_idx, column=10, value=item.get('certificateNumber', '')).border = border
        ws.cell(row=row_idx, column=11, value=item.get('currentLocation', '')).border = border
        ws.cell(row=row_idx, column=12, value=item.get('currentProjectName', '')).border = border
        ws.cell(row=row_idx, column=13, value=item.get('usageTurbineId', '')).border = border
        ws.cell(row=row_idx, column=14, value=item.get('usageConnection', '')).border = border
        ws.cell(row=row_idx, column=15, value=item.get('usageOperation', '')).border = border
        ws.cell(row=row_idx, column=16, value=item.get('usageUser', '')).border = border
        ws.cell(row=row_idx, column=17, value=item.get('usageDate', '')).border = border
        ws.cell(row=row_idx, column=18, value=item.get('notes', '')).border = border

    print(f"✓ Equipamentos sheet created with {len(equipamentos_data)} items")

# ═════════════════════════════════════════════════════════════════
# FUNÇÃO PRINCIPAL
# ═════════════════════════════════════════════════════════════════

def generate_excel_report(project_name, data_by_phase, selected_phases, output_path, language='pt', complete_report=True):
    """
    Gera relatório Excel com dados de instalação
    """
    print(f"\n{'='*60}")
    print(f"🚀 INICIANDO GERAÇÃO DE RELATÓRIO EXCEL - VERSÃO 3")
    print(f"{'='*60}")
    print(f"📁 Projeto: {project_name}")
    print(f"🌍 Idioma: {language.upper()}")
    print(f"📊 Fases selecionadas: {len(selected_phases)}")
    print(f"📋 Modo: {'Completo (com capa/resumo/dashboard)' if complete_report else 'Apenas fases'}")
    
    wb = Workbook()
    
    # Remover sheet padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    if complete_report:
        # FASE 1: CAPA E RESUMO EXECUTIVO
        _create_cover_sheet(wb, project_name, language)
        _create_executive_summary_v2(wb, project_name, data_by_phase, language)
        
        # FASE 2: DASHBOARD COM GRÁFICOS
        _create_dashboard_v3(wb, project_name, data_by_phase, language)
    
    # FASE 3: ANÁLISES AVANÇADAS
    gruas_pads_data = data_by_phase.get('gruasPads', [])
    gruas_gerais_data = data_by_phase.get('gruasGerais', [])
    equipamentos_data = data_by_phase.get('equipamentos', [])
    ncrs_data = data_by_phase.get('ncrs', [])
    
    if (gruas_pads_data or gruas_gerais_data) or complete_report:
        _add_deviation_analysis_sheet(wb, data_by_phase, language)
        _add_crane_analysis_sheet(wb, gruas_pads_data, gruas_gerais_data, language)
        _add_critical_observations_sheet(wb, data_by_phase, language)
    
    # FASE 4: SHEETS DAS FASES
    current_sheet_idx = 3 if complete_report else 0
    
    # Processar Receção
    if 'recepcao' in selected_phases:
        recepcao_data = data_by_phase.get('recepcao', [])
        if recepcao_data:
            print(f"[OK] Adding 'Receção' sheet with {len(recepcao_data)} items")
            _add_recepcao_sheet(wb, recepcao_data, language)
            sheet_name = 'Receção' if language == 'pt' else 'Reception'
            ws = wb[sheet_name]
            _add_project_header(ws, project_name, sheet_name, language)
            _add_auto_filters(ws, 4, 7)  # Linha 1 + 3 linhas inseridas = linha 4
            _add_footer(ws, language)
            current_sheet_idx += 1
    
    # Processar Preparação
    if 'preparacao' in selected_phases:
        preparacao_data = data_by_phase.get('preparacao', [])
        if preparacao_data:
            print(f"[OK] Adding 'Preparação' sheet with {len(preparacao_data)} items")
            _add_preparacao_sheet(wb, preparacao_data, language)
            sheet_name = 'Preparação' if language == 'pt' else 'Preparation'
            ws = wb[sheet_name]
            _add_project_header(ws, project_name, sheet_name, language)
            _add_auto_filters(ws, 4, 9)  # Linha 1 + 3 linhas inseridas = linha 4
            _add_footer(ws, language)
            current_sheet_idx += 1
    
    # Processar Pré-Assemblagem
    if 'preAssemblagem' in selected_phases:
        pre_assemblagem_data = data_by_phase.get('preAssemblagem', [])
        if pre_assemblagem_data:
            print(f"[OK] Adding 'Pré-Assemblagem' sheet with {len(pre_assemblagem_data)} items")
            _add_pre_assemblagem_sheet(wb, pre_assemblagem_data, language)
            sheet_name = 'Pré-Assemblagem' if language == 'pt' else 'Pre-Assembly'
            ws = wb[sheet_name]
            _add_project_header(ws, project_name, sheet_name, language)
            _add_auto_filters(ws, 4, 9)  # Linha 1 + 3 linhas inseridas = linha 4
            _add_footer(ws, language)
            current_sheet_idx += 1
    
    # Processar Assemblagem
    if 'assemblagem' in selected_phases:
        assemblagem_data = data_by_phase.get('assemblagem', [])
        if assemblagem_data:
            print(f"[OK] Adding 'Assemblagem' sheet with {len(assemblagem_data)} items")
            _add_assemblagem_sheet(wb, assemblagem_data, language)
            sheet_name = 'Assemblagem' if language == 'pt' else 'Assembly'
            ws = wb[sheet_name]
            _add_project_header(ws, project_name, sheet_name, language)
            _add_auto_filters(ws, 4, 9)  # Linha 1 + 3 linhas inseridas = linha 4
            _add_footer(ws, language)
            current_sheet_idx += 1
    
    # Processar Torque & Tensioning
    if 'torqueTensionamento' in selected_phases:
        torque_data = data_by_phase.get('torqueTensionamento', [])
        if torque_data:
            print(f"[OK] Adding 'Torque' sheet with {len(torque_data)} items")
            _add_torque_sheet(wb, torque_data, language)
            sheet_name = 'Torque & Tensionamento' if language == 'pt' else 'Torque & Tensioning'
            ws = wb[sheet_name]
            _add_project_header(ws, project_name, sheet_name, language)
            _add_auto_filters(ws, 4, 8)  # Linha 1 + 3 linhas inseridas = linha 4
            _add_footer(ws, language)
            current_sheet_idx += 1
    
    # Processar Fases Finais
    if 'fasesFinal' in selected_phases:
        fases_data = data_by_phase.get('fasesFinal', [])
        if fases_data:
            print(f"[OK] Adding 'Fases Finais' sheet with {len(fases_data)} items")
            _add_fases_finais_sheet(wb, fases_data, language)
            sheet_name = 'Fases Finais' if language == 'pt' else 'Final Phases'
            ws = wb[sheet_name]
            _add_project_header(ws, project_name, sheet_name, language)
            _add_auto_filters(ws, 4, 9)  # Linha 1 + 3 linhas inseridas = linha 4
            _add_footer(ws, language)
            current_sheet_idx += 1
    
    # Processar Gruas Pads
    if gruas_pads_data:
        print(f"[OK] Adding 'Gruas - Pads' sheet with {len(gruas_pads_data)} items")
        _add_gruas_pads_sheet(wb, gruas_pads_data, language)
        ws = wb['Gruas - Pads']
        _add_project_header(ws, project_name, 'Gruas - Pads', language)
        _add_auto_filters(ws, 4, 12)  # Linha 1 + 3 linhas inseridas = linha 4
        _add_footer(ws, language)
        current_sheet_idx += 1
    
    # Processar Gruas Gerais
    if gruas_gerais_data:
        print(f"[OK] Adding 'Gruas - Gerais' sheet with {len(gruas_gerais_data)} items")
        _add_gruas_gerais_sheet(wb, gruas_gerais_data, language)
        ws = wb['Gruas - Gerais']
        _add_project_header(ws, project_name, 'Gruas - Gerais', language)
        _add_auto_filters(ws, 4, 12)  # Linha 1 + 3 linhas inseridas = linha 4
        _add_footer(ws, language)

    # Processar Equipamentos
    if equipamentos_data:
        print(f"[OK] Adding 'Equipamentos' sheet with {len(equipamentos_data)} items")
        _add_equipamentos_sheet(wb, equipamentos_data, language)
        sheet_name = 'Equipamentos' if language == 'pt' else 'Equipment'
        ws = wb[sheet_name]
        _add_project_header(ws, project_name, sheet_name, language)
        _add_auto_filters(ws, 4, 18)  # Linha 1 + 3 linhas inseridas = linha 4
        _add_footer(ws, language)

    if ncrs_data:
        print(f"[OK] Adding 'NCRs' sheet with {len(ncrs_data)} items")
        _add_ncr_sheet(wb, ncrs_data, language)
        ws = wb['NCRs']
        _add_project_header(ws, project_name, 'NCRs', language)
        _add_footer(ws, language)
    
    # Garantir que o workbook tem pelo menos uma sheet visível
    if not wb.sheetnames:
        empty_sheet_name = 'Sem Dados' if language == 'pt' else 'No Data'
        ws = wb.create_sheet(empty_sheet_name)
        ws['A1'] = 'Não existem dados para as fases selecionadas.' if language == 'pt' else 'There is no data for the selected phases.'
        ws['A1'].font = Font(name='Arial', size=11, color=COLORS['dark_gray'])
        ws.column_dimensions['A'].width = 60

    # 📝 SALVAR
    print(f"\n{'─'*60}")
    print(f"💾 Salvando relatório em: {output_path}")
    wb.save(output_path)
    print(f"✅ RELATÓRIO GERADO COM SUCESSO!")
    print(f"{'='*60}\n")

# ═════════════════════════════════════════════════════════════════
# PONTO DE ENTRADA
# ═════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    input_data = json.loads(sys.stdin.read())
    
    project_name = input_data['projectName']
    data_by_phase = input_data['dataByPhase']
    selected_phases = input_data['selectedPhases']
    output_path = input_data['outputPath']
    language = input_data.get('language', 'pt')
    complete_report = input_data.get('completeReport', True)
    
    generate_excel_report(project_name, data_by_phase, selected_phases, output_path, language, complete_report)

